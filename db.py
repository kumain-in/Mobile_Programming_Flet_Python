"""
db.py -- the "database" layer for Manajemen Mahasiswa.

Everything that reads or writes student data goes through this module; the UI
(main.py) never opens the spreadsheet itself. The store is a single Excel file,
data/students.xlsx (one sheet, "Mahasiswa"), which is auto-created and seeded
with 24 sample rows the first time the app runs.

WHY EXCEL: it needs no server and the file can be opened/inspected in Excel.
CAVEAT: there is no locking or transactions, so this suits a small single-user
or light internal app. To scale, swap the bodies of the functions below for
SQLite/Postgres -- the function signatures (get_all/add/update/delete) can stay
the same, so main.py would not need to change.
"""

import openpyxl
from pathlib import Path
from dataclasses import dataclass
from typing import List

# --- where the data lives + the sheet layout ---
DATA_FILE = Path(__file__).parent / "data" / "students.xlsx"
SHEET_NAME = "Mahasiswa"
HEADERS = ["NIM", "Nama", "Prodi", "Angkatan", "Email", "IPK", "Status"]


@dataclass
class Student:
    """One student record. Mirrors one row of the spreadsheet (minus the header).
    `status` is one of: aktif | cuti | do | lulus (see constants.STATUS)."""
    nim: str          # student number; the primary key (unique, never changed)
    nama: str
    prodi: str        # program / major
    angkatan: int     # entry year
    email: str
    ipk: float        # GPA, 0.00-4.00
    status: str


def _email_for(nama: str) -> str:
    """Build a campus email from a name (used only to seed sample data),
    e.g. 'Budi Santoso' -> 'budi.santoso@student.univ.ac.id'."""
    parts = [p for p in "".join(
        c for c in nama.lower() if c.isalpha() or c.isspace()).split() if p]
    return ".".join(parts[:2]) + "@student.univ.ac.id"


# Sample roster written on first run: (nim, nama, prodi, angkatan, ipk, status)
_SEED_ROWS = [
    ("2021103001", "Rangga Aditya Pratama", "Teknik Informatika",        2021, 3.62, "aktif"),
    ("2021103002", "Siti Nurhaliza",        "Sistem Informasi",          2021, 3.81, "aktif"),
    ("2020201014", "Budi Santoso",          "Manajemen",                 2020, 3.24, "lulus"),
    ("2022104027", "Dewi Lestari",          "Psikologi",                 2022, 3.74, "aktif"),
    ("2023105009", "Ahmad Fauzi",           "Ilmu Hukum",                2023, 2.91, "aktif"),
    ("2021103045", "Putri Maharani",        "Teknik Informatika",        2021, 3.55, "cuti"),
    ("2019301002", "Hendra Gunawan",        "Teknik Sipil",              2019, 2.48, "do"),
    ("2022202018", "Anisa Rahmawati",       "Akuntansi",                 2022, 3.93, "aktif"),
    ("2023106033", "Bagus Wicaksono",       "Ilmu Komunikasi",           2023, 3.18, "aktif"),
    ("2020103077", "Citra Kirana",          "Sistem Informasi",          2020, 3.67, "lulus"),
    ("2024107005", "Dimas Prasetyo",        "Desain Komunikasi Visual",  2024, 3.40, "aktif"),
    ("2022108011", "Eka Putri Wulandari",   "Arsitektur",                2022, 3.71, "aktif"),
    ("2021109050", "Fajar Nugroho",         "Teknik Elektro",            2021, 2.76, "cuti"),
    ("2023105021", "Gita Savitri",          "Ilmu Hukum",                2023, 3.88, "aktif"),
    ("2020110004", "Indah Permatasari",     "Kedokteran",                2020, 3.96, "aktif"),
    ("2019103088", "Joko Susilo",           "Teknik Informatika",        2019, 3.05, "lulus"),
    ("2022202044", "Kartika Sari",          "Akuntansi",                 2022, 3.49, "aktif"),
    ("2024106019", "Lukman Hakim",          "Ilmu Komunikasi",           2024, 2.63, "aktif"),
    ("2021104031", "Maya Anggraini",        "Psikologi",                 2021, 3.59, "cuti"),
    ("2023103012", "Nanda Pratama",         "Teknik Informatika",        2023, 3.33, "aktif"),
    ("2020201066", "Oki Setiana Putri",     "Manajemen",                 2020, 3.78, "lulus"),
    ("2022105007", "Rizki Ramadhan",        "Ilmu Hukum",                2022, 2.39, "do"),
    ("2024108002", "Salsa Wijaya",          "Arsitektur",                2024, 3.52, "aktif"),
    ("2021110023", "Tari Handayani",        "Kedokteran",                2021, 3.85, "aktif"),
]


def _seed() -> List[Student]:
    """Turn the raw seed tuples into Student objects (email derived from name)."""
    return [
        Student(nim, nama, prodi, ang, _email_for(nama), ipk, status)
        for (nim, nama, prodi, ang, ipk, status) in _SEED_ROWS
    ]


def _ensure_file() -> None:
    """Create data/students.xlsx with a header row + seed data if it doesn't
    exist yet. Safe to call on every access -- it's a no-op once the file exists."""
    if DATA_FILE.exists():
        return
    DATA_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(HEADERS)
    for s in _seed():
        ws.append([s.nim, s.nama, s.prodi, s.angkatan, s.email, s.ipk, s.status])
    wb.save(DATA_FILE)


def _load():
    """Open the workbook (creating it first if needed) and return (wb, sheet)."""
    _ensure_file()
    wb = openpyxl.load_workbook(DATA_FILE)
    return wb, wb[SHEET_NAME]


def get_all() -> List[Student]:
    """Return every student, newest-first (rows are reversed so the most
    recently added appears at the top of the list, matching the prototype).
    Cells are coerced to the right types (nim->str, angkatan->int, ipk->float)."""
    _, ws = _load()
    rows: List[Student] = []
    for r in ws.iter_rows(min_row=2, values_only=True):   # skip the header row
        if r[0] is None:                                  # skip blank rows
            continue
        rows.append(Student(
            nim=str(r[0]),
            nama=str(r[1]),
            prodi=str(r[2]),
            angkatan=int(r[3]),
            email=str(r[4]) if r[4] else "",
            ipk=round(float(r[5]), 2),
            status=str(r[6]),
        ))
    rows.reverse()
    return rows


def exists(nim: str) -> bool:
    """True if a student with this NIM is already stored."""
    return any(s.nim == nim for s in get_all())


def add(student: Student) -> Student:
    """Append a new student as the last row of the sheet and save."""
    wb, ws = _load()
    ws.append([student.nim, student.nama, student.prodi, student.angkatan,
               student.email, student.ipk, student.status])
    wb.save(DATA_FILE)
    return student


def update(updated: Student) -> None:
    """Find the row whose NIM matches and overwrite all its columns in place.
    (NIM is the key and is locked in the UI, so it is never changed here.)"""
    wb, ws = _load()
    for row in ws.iter_rows(min_row=2):
        if str(row[0].value) == updated.nim:
            (row[0].value, row[1].value, row[2].value, row[3].value,
             row[4].value, row[5].value, row[6].value) = (
                updated.nim, updated.nama, updated.prodi, updated.angkatan,
                updated.email, updated.ipk, updated.status)
            break
    wb.save(DATA_FILE)


def delete(nim: str) -> None:
    """Remove the row whose NIM matches and save."""
    wb, ws = _load()
    for row in ws.iter_rows(min_row=2):
        if str(row[0].value) == nim:
            ws.delete_rows(row[0].row)
            break
    wb.save(DATA_FILE)
